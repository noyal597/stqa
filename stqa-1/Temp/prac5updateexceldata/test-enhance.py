from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl
from time import sleep

def test():
    try:
        # Set up the webdriver
        service = Service("C:\selenium\chromedriver-win64\chromedriver.exe")
        driver = webdriver.Chrome(service=service)
        
        # Navigate to the HTML page
        driver.get("http://127.0.0.1:5000/")
        sleep(2)
        
        # Create a list of student data to be entered
        students = [["student-1", "17", "F"], ["student-2", "12", "F"], ["student-3", "13", "F"], ["student-4", "14", "F"], ["student-5", "15","F"]]
        
        # Create an empty list to store the student data
        student_data = []

        for student in students:
            # Locate the input fields on the page
            name_input = driver.find_element(By.ID,"name")
            age_input = driver.find_element(By.ID,"age")
            grade_input = driver.find_element(By.ID,"grade")

            # Enter data into the input fields
            name_input.send_keys(student[0])
            age_input.send_keys(student[1])
            grade_input.send_keys(student[2])

            # Submit the form
            grade_input.send_keys(Keys.RETURN)
            sleep(3)
            msg = driver.find_element(By.ID,"msg").text
            if msg == "Student record updated successfully!":
                print(f"Record for {student[0]} updated successfully!")
            else:
              print("A wrong input was given aborting updates")
              return

        # Locate the table containing the student records
        table = driver.find_element(By.ID,"myTable")

        # Get the rows of the table
        rows = table.find_elements(By.TAG_NAME,"tr")

        # Loop through each row of the table and extract the relevant data
        for row in rows[1:]: # Skip the first row (header row)
            cells = row.find_elements(By.TAG_NAME,"td")
            name = cells[0].text
            age = cells[1].text
            grade = cells[2].text

            # Append the data to the student_data list
            student_data.append([name, age, grade])

        # Close the webdriver
        driver.quit()
        # Load or create an Excel file containing the student records
        print("startng to update data in excel sheet")
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write the headers to the first row of the Excel file
        ws.cell(row=1, column=1).value = "Name"
        ws.cell(row=1, column=2).value = "Age"
        ws.cell(row=1, column=3).value = "Grade"

        # Write the student data to the Excel file
        for i, data in enumerate(student_data):
            ws.cell(row=i+2, column=1).value = data[0] # Write name (skip the first row (header row))
            ws.cell(row=i+2, column=2).value = data[1] # Write age
            ws.cell(row=i+2, column=3).value = data[2] # Write grade

        # Save the Excel file
        wb.save("prac5.xlsx")
        print("data successfully added to sheet")
            
    except Exception as e:
        print(e)

test()